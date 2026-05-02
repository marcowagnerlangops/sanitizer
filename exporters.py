# exporters.py

from __future__ import annotations

from collections import Counter
from io import BytesIO
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sanitizer_core import SegmentRecord


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
OK_FILL = PatternFill(fill_type="solid", fgColor="D9EAD3")
MINOR_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
MAJOR_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")
CRITICAL_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")


COLUMNS = [
    "Record ID",
    "File",
    "Type",
    "Unit ID",
    "Source Lang",
    "Target Lang",
    "Source",
    "Target",
    "Severity",
    "Issue Count",
    "Issue Categories",
    "Issue Details",
    "LQA Severity",
    "LQA Penalty",
    "LQA Details",
    "Repair Actions",
    "Notes",
]


def _style(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"

    # LQA Severity column = 13
    for row in range(2, ws.max_row + 1):
        sev = ws.cell(row=row, column=13).value or "OK"

        if sev == "Critical":
            fill = CRITICAL_FILL
        elif sev == "Major":
            fill = MAJOR_FILL
        elif sev == "Minor":
            fill = MINOR_FILL
        else:
            fill = OK_FILL

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max_len + 2, 80)


def _append_record(ws, r: SegmentRecord):
    ws.append([
        r.record_id,
        r.file_name,
        r.file_type,
        r.unit_id,
        r.source_lang,
        r.target_lang,
        r.source_text,
        r.target_text,
        r.severity,
        r.issue_count,
        r.issue_categories,
        r.issue_details,
        r.lqa_severity,
        r.lqa_penalty,
        r.lqa_details,
        r.repair_actions,
        r.notes,
    ])


def build_xlsx_report(records: List[SegmentRecord], stats: Dict[str, object]) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Segments"
    ws.append(COLUMNS)
    for r in records:
        _append_record(ws, r)
    _style(ws)

    issues = wb.create_sheet("Issues Only")
    issues.append(COLUMNS)
    for r in records:
        if r.issue_count:
            _append_record(issues, r)
    _style(issues)

    critical_ws = wb.create_sheet("Critical Major")
    critical_ws.append(COLUMNS)
    for r in records:
        if r.lqa_severity in {"Critical", "Major"}:
            _append_record(critical_ws, r)
    _style(critical_ws)

    repairs = wb.create_sheet("Repair Actions")
    repairs.append(COLUMNS)
    for r in records:
        if r.repair_actions:
            _append_record(repairs, r)
    _style(repairs)

    lqa_ws = wb.create_sheet("LQA Summary")
    lqa_ws.append(["Metric", "Value"])
    lqa_ws.append(["Quality Score", stats.get("quality_score", 100)])
    lqa_ws.append(["Quality Label", stats.get("quality_label", "Excellent")])
    lqa_ws.append(["Total LQA Penalty", stats.get("total_lqa_penalty", 0)])
    lqa_ws.append(["Critical Issues", stats.get("critical_issues", 0)])
    lqa_ws.append(["Major Issues", stats.get("major_issues", 0)])
    lqa_ws.append(["Minor Issues", stats.get("minor_issues", 0)])
    lqa_ws.append([])
    lqa_ws.append(["Segment Severity", "Count"])

    for key, value in stats.get("lqa_segment_severity", {}).items():
        lqa_ws.append([key, value])

    _style(lqa_ws)

    stats_ws = wb.create_sheet("Statistics")
    stats_ws.append(["Metric", "Value"])

    for key, value in stats.items():
        if isinstance(value, (dict, Counter)):
            stats_ws.append([key, ""])
            for sub_key, sub_value in value.items():
                stats_ws.append([f"  {sub_key}", sub_value])
        else:
            stats_ws.append([key, value])

    _style(stats_ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
